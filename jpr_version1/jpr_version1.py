import sys
import re
import openpyxl
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QAction, QTableWidget, QTableWidgetItem, QVBoxLayout, \
QPushButton, QHBoxLayout, QGroupBox, QDialog, QFileDialog, QAction
from PyQt5.QtMultimedia import QMediaPlaylist, QMediaPlayer, QMediaContent
from PyQt5.QtGui import QIcon, QColor
from PyQt5.QtCore import pyqtSlot, QSize, QUrl
 
class App(QMainWindow):
 
    def __init__(self):
        super().__init__()
        self.player = QMediaPlayer()
        self.playlist = QMediaPlaylist()
        self.playlist.setPlaybackMode(QMediaPlaylist.Sequential)
        self.title = 'JPR Reader GUI version1'
        self.left = 50
        self.top = 100
        self.width = 1050
        self.height = 550
        self.fileReady = False
        self.tableRow = 5
        self.tableCol = 10
        self.row = 2
        self.words = []
        self.couple = []
        self.audiolist = []
        self.initUI()
 
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        # menu
        menubar = self.menuBar()
        filemenu = menubar.addMenu('File')
        fileAct = QAction('Open File', self)
        fileAct.setShortcut('Ctrl+O')
        filemenu.addAction(fileAct)
        fileAct.triggered.connect(self.openFileNameDialog)

        # status_bar
        self.statusbar = self.statusBar()
 
        # gui object
        self.wid = QWidget(self)
        self.setCentralWidget(self.wid)
        self.createHL()
        self.createTable()
 
        # layout
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.horizontalGroupBox) 
        self.layout.addWidget(self.tableWidget) 
        self.wid.setLayout(self.layout) 
 
        # Show widget
        self.show()

    def openFileNameDialog(self):    
        fileName, _ = QFileDialog.getOpenFileName(self,"Open file", "", "XML files (*.xml, *.xlsx)")
        if fileName:
            self.wb = openpyxl.load_workbook(fileName.strip())
            self.ws = self.wb.active
            self.fileReady = True
            self.statusbar.showMessage('succeed to load file...')
            self.row = 2
            self.tableWidget.clearContents()
        else:
            self.statusbar.showMessage('Fail to load file...')
        

    def createHL(self):
        self.horizontalGroupBox = QGroupBox("Controller")
        layout = QHBoxLayout()
 
        pre_button = QPushButton('', self)
        pre_button.clicked.connect(self.pre_click)
        pre_button.setIcon(QIcon('.\\img\\pre.png'))
        pre_button.setIconSize(QSize(250,250))
        layout.addWidget(pre_button) 
 
        cur_button = QPushButton('', self)
        cur_button.clicked.connect(self.cur_click)
        cur_button.setIcon(QIcon('.\\img\\cur.png'))
        cur_button.setIconSize(QSize(250,250))
        layout.addWidget(cur_button) 
 
        next_button = QPushButton('', self)
        next_button.clicked.connect(self.next_click)
        next_button.setIcon(QIcon('.\\img\\next.png'))
        next_button.setIconSize(QSize(250,250))
        layout.addWidget(next_button) 
 
        self.horizontalGroupBox.setLayout(layout)
 
    def createTable(self):
       # Create table
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(self.tableRow)
        self.tableWidget.setColumnCount(self.tableCol)

        # Horizontal header
        self.tableWidget.setHorizontalHeaderItem(0, QTableWidgetItem("포장순번"))
        self.tableWidget.setHorizontalHeaderItem(1, QTableWidgetItem("거래처명"))
        self.tableWidget.setHorizontalHeaderItem(2, QTableWidgetItem("배송센터"))
        self.tableWidget.setHorizontalHeaderItem(3, QTableWidgetItem("박스"))
        self.tableWidget.setHorizontalHeaderItem(4, QTableWidgetItem("깔개"))
        self.tableWidget.setHorizontalHeaderItem(5, QTableWidgetItem("상부"))
        self.tableWidget.setHorizontalHeaderItem(6, QTableWidgetItem("하부"))
        self.tableWidget.setHorizontalHeaderItem(7, QTableWidgetItem("행잉"))
        self.tableWidget.setHorizontalHeaderItem(8, QTableWidgetItem("평대"))
        self.tableWidget.setHorizontalHeaderItem(9, QTableWidgetItem("배너"))

        self.tableWidget.move(0,0)

    def setTable(self):
        # shifting other rows
        for r in range(1, self.tableRow):
            for c in range(self.tableCol):
                try:
                    self.tableWidget.item(r,c).setBackground(QColor(255,255,255))
                    self.tableWidget.setItem(r-1 , c, self.tableWidget.item(r,c).clone())
                except:
                    pass

        # set current row
        self.tableWidget.setItem(self.tableRow - 1, 0, QTableWidgetItem(self.num))
        self.tableWidget.setItem(self.tableRow - 1, 1, QTableWidgetItem(self.partner))
        self.tableWidget.setItem(self.tableRow - 1, 2, QTableWidgetItem(self.parcel))
        self.tableWidget.setItem(self.tableRow - 1, 3, QTableWidgetItem(self.box))
        self.tableWidget.setItem(self.tableRow - 1, 4, QTableWidgetItem(self.kkar))
        self.tableWidget.setItem(self.tableRow - 1, 5, QTableWidgetItem(self.sang))
        self.tableWidget.setItem(self.tableRow - 1, 6, QTableWidgetItem(self.ha))
        self.tableWidget.setItem(self.tableRow - 1, 7, QTableWidgetItem(self.hang))
        self.tableWidget.setItem(self.tableRow - 1, 8, QTableWidgetItem(self.pyeng))
        self.tableWidget.setItem(self.tableRow - 1, 9, QTableWidgetItem(self.ban))

        for c in range(self.tableCol):
            self.tableWidget.item(self.tableRow - 1, c).setBackground(QColor(255,255,0))

    def read(self):
        if not self.fileReady:
            self.openFileNameDialog()

        #포장 순번
        self.num = str(self.ws.cell(row = self.row, column = 3).value[2:]).strip()
        print('<<<', self.num, '>>>', end = ' / ')
        self.couple.append('포장순번')
        self.parsing(self.num)

        #거래처명
        self.partner = str(self.ws.cell(row = self.row, column = 5).value).strip()
        print('거래처:', self.partner, end = ' / ')

        #택배발송
        self.parcel = str(self.ws.cell(row = self.row, column = 4).value).strip()
        print('배송센터:', self.parcel, end = ' / ')
        self.couple.append('배송센터')
        self.parsing(self.parcel)

        #박스
        self.box = str(self.ws.cell(row = self.row, column = 2).value).strip()
        print('박스:', self.box, end =' / ')
        self.couple.append('박스')
        self.parsing(self.box)


        #깔개
        self.kkar = str(self.ws.cell(row = self.row, column = 7).value).strip()
        print('깔개:', self.kkar, end =' / ')
        self.couple.append('깔개')
        self.parsing(self.kkar)


        #상부
        self.sang = str(self.ws.cell(row = self.row, column = 8).value).strip()
        print('상부:', self.sang, end =' / ')
        self.couple.append('상부')
        self.parsing(self.sang)


        #하부
        self.ha = str(self.ws.cell(row = self.row, column = 9).value).strip()
        print('하부:', self.ha, end =' / ')
        self.couple.append('하부')
        self.parsing(self.ha)


        #행잉
        self.hang = str(self.ws.cell(row = self.row, column = 10).value).strip()
        print('행잉:', self.hang, end =' / ')
        self.couple.append('행잉')
        self.parsing(self.hang)


        #평대
        self.pyeng = str(self.ws.cell(row = self.row, column = 11).value).strip()
        print('평대:', self.pyeng, end =' / ')
        self.couple.append('평대')
        self.parsing(self.pyeng)


        #배너
        self.ban = str(self.ws.cell(row = self.row, column = 12).value).strip()
        print('배너:', self.ban)
        self.couple.append('배너')
        self.parsing(self.ban)


        self.setTable()
        print(self.words)

    def parsing(self, val):
        if val == 'None' or val == '':
            self.couple.append('없음')

        else:
            if '(' in val or '/' in val:
                arr = re.split('[(/]',val)
                for i in range(len(arr)):
                    if ')' in arr[i]:
                        arr[i] = arr[i][:arr[i].index(')')]
                self.couple.append(arr)

            else:
                self.couple.append(val)

        self.words.append(self.couple)
        self.couple = [] 

    def load_audiolist(self):
        for item, value in self.words:
            if item == '배송센터':
                if value == '택배발송':
                    self.audiolist.append('택배발송')
            else:
                self.audiolist.append(item.strip())
                # 포장순번
                if item == '포장순번':
                    for i in range(3):
                        self.audiolist.append('_' + value[i])

                # in / or () case
                elif isinstance(value, list):
                    for name in value:
                        # '2개' case
                        if '2' in name:
                            self.audiolist.append('2')
                        else:
                            self.audiolist.append(name.strip())
                else:
                    self.audiolist.append(value.strip())

            # beep
            self.audiolist.append('beep')

    def speak(self):
        self.playlist.clear()
        for clip in self.audiolist:
            url = QUrl.fromLocalFile('.\\audio_clips\\' + clip + '.mp3')
            #print(url)
            self.playlist.addMedia(QMediaContent(url))
        self.player.setPlaylist(self.playlist)
        self.player.play()
            

    @pyqtSlot()
    def pre_click(self):
        if not self.fileReady:
            pass
        
        else:
            if self.row == 2:
                self.statusbar.showMessage("Can't be previous.")
            else:
                self.row -= 1

        del self.words[:]
        del self.audiolist[:]
        self.read()
        self.load_audiolist()
        self.speak()


    @pyqtSlot()
    def cur_click(self):
        del self.words[:]
        del self.audiolist[:]
        self.read()
        self.load_audiolist()
        self.speak()


    @pyqtSlot()
    def next_click(self):
        if not self.fileReady:
            pass

        else:
            if self.row == self.ws.max_row:
                self.statusbar.showMessage("It's over.")
            else:
                self.row += 1

        del self.words[:]
        del self.audiolist[:]
        self.read()
        self.load_audiolist()
        self.speak()      
        
 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())