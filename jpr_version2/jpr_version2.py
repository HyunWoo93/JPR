import sys
import re
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QAction, QTableWidget, QTableWidgetItem, QVBoxLayout, \
QPushButton, QHBoxLayout, QGroupBox, QDialog, QFileDialog, QAction, QTabWidget, QLabel, QTableView, QAbstractItemView, \
QMessageBox
from PyQt5.QtMultimedia import QMediaPlaylist, QMediaPlayer, QMediaContent
from PyQt5.QtGui import QIcon, QColor
from PyQt5.QtCore import pyqtSlot, QSize, QUrl, Qt
from shutil import copyfile
from os import listdir, remove
 
class App(QMainWindow):
 
    def __init__(self):
        super().__init__()
        self.player = QMediaPlayer()
        self.playlist = QMediaPlaylist()
        self.playlist.setPlaybackMode(QMediaPlaylist.Sequential)
        self.title = 'JPR Reader GUI version2'
        self.left = 50
        self.top = 100
        self.width = 1200
        self.height = 800
        self.fileReady = False
        self.tableRow = 5
        self.tableCol = 15
        self.row = 2
        self.audiolist = []
        self.configFile = ".//configurationFile.xlsx"
        self.dict = {'num': None, 'partner': None, 'parcel': None, 'exception': None, 'box': None}
        self.initUI()
 
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        # menu
        menubar = self.menuBar()
        filemenu = menubar.addMenu('File')
        fileAct = QAction('Open File', self)
        fileAct.setShortcut('Ctrl+O')
        filemenu.addAction(fileAct)
        fileAct.triggered.connect(self.openFileNameDialog)

        # status_bar
        self.statusbar = self.statusBar()
 
        # tabs
        self.tabs = QTabWidget()
        self.tab1 = QWidget(self)
        self.tab2 = QWidget(self)
        self.tabs.addTab(self.tab1,"조작")
        self.tabs.addTab(self.tab2,"설정")
        self.setCentralWidget(self.tabs)

        # tab1 gui
        self.ButtonGroupBox = self.createButtonGroupBox()
        self.createLogTable()
        self.tab1.layout = QVBoxLayout()
        self.tab1.layout.addWidget(self.ButtonGroupBox) 
        self.tab1.layout.addWidget(self.logTable) 
        self.tab1.setLayout(self.tab1.layout)

        # tab2 gui
        self.explanation = QLabel()
        self.explanation.setText(
            """<<< JPR reader 설정 테이블 >>>    
            이곳에서 JPR reader가 읽어주는 항목과 아이템을 설정할 수 있습니다.
            항목과 아이템의 설정 방법은 셀을 더블 클릭한 후 이름을 입력하는 방식으로진행됩니다.
            그 후 떠오른 파일 탐색기에서 지정할 mp3파일을 선택해 주세요.
            설정의 확인은 셀의 색으로 확인 가능합니다. 지정이 완료된 항목은 노란색, 아이템은 하늘색으로 표시됩니다.
            지정이 되지 않은 셀은 빨간색으로 표시됩니다. 행과 열 버튼으로 테이블의 크기를 조절할 수 있으며
            초기화시 모든 설정은 사라지며 테이블은 5by5로 초기화 됩니다.

            <<<주의사항>>>
            1. 첫번째 열은 반드시 항목을 입력해 주세요.
            2. 입력되는 이름은 엑셀파일에서 표현된 명칭과 완벽히 일치해야 합니다.(엔터나 스페이스, 오타 주의!)
            3. 초기화를 누르면 처음부터 모든 항목과 아이템을 지정해야 합니다.
            4. 엑셀 파일의 ()나 /을 통해 구분한 아이템은 따로 입력해 주세요.
            5. 사용하는 엑셀파일의 구조를 유지해 주세요. 변경시 프로그램의 수정이 필요할 수 있습니다.(예: '박스'항목은 항상 있을 것으로 간주됩니다.)

            제작자 정보: 김현우(gguussddnn@gmail.com)""")
        self.explanation.setAlignment(Qt.AlignCenter)
        self.createConfigTable()
        self.createHorizontalButtons2()
        self.tab2.layout = QVBoxLayout()
        self.tab2.layout.addWidget(self.explanation) 
        self.tab2.layout.addWidget(self.configTable)
        self.tab2.layout.addWidget(self.horizontalButtons2) 
        self.tab2.setLayout(self.tab2.layout)

 
        # Show widget
        self.show()

    def openFileNameDialog(self):    
        fileName, _ = QFileDialog.getOpenFileName(self,"Open file", "", "All Files (*)")
        if not fileName:
            self.statusbar.showMessage('Fail to load file...')

        else:
            self.wb = load_workbook(fileName.strip())
            self.ws = self.wb.active
            self.fileReady = True
            self.row = 2
            self.statusbar.showMessage('succeed to load file...')
            self.logTable.clear()
            
            # set logTable's horizontal header
            self.logTable.setHorizontalHeaderItem(0, QTableWidgetItem("포장순번"))
            self.logTable.setHorizontalHeaderItem(1, QTableWidgetItem("거래처명"))
            self.logTable.setHorizontalHeaderItem(2, QTableWidgetItem("배송센터"))
            self.logTable.setHorizontalHeaderItem(3, QTableWidgetItem("특이사항"))
            self.logTable.setHorizontalHeaderItem(4, QTableWidgetItem("박스"))

            # initialize dictionary
            self.dict = {'num': None, 'partner': None, 'parcel': None, 'exception': None, 'box': None}
            col = 6
            num = 0
            header = str(self.ws.cell(row = 1, column = col + 1).value).strip()
            while header != 'None':
                self.dict[header] = None
                col = col + 1
                num = num + 1
                header = str(self.ws.cell(row = 1, column = col + 1).value).strip()

            self.tableCol = 5 + num
            self.logTable.setColumnCount(self.tableCol)

            for c in range(5, self.tableCol):
                self.logTable.setHorizontalHeaderItem(c, QTableWidgetItem(list(self.dict.keys())[c]))
            
    def createButtonGroupBox(self):
        buttonGroupBox = QGroupBox("Controller")
        vLayout = QVBoxLayout()

        pre_button = QPushButton('w', self)
        pre_button.clicked.connect(self.pre_click)
        pre_button.setIcon(QIcon('.\\img\\up-arrow.png'))
        pre_button.setIconSize(QSize(600,100))
        pre_button.setShortcut('w')
        vLayout.addWidget(pre_button) 

        hBottensWidget = self.createHButtons()
        vLayout.addWidget(hBottensWidget)

        next_button = QPushButton('x', self)
        next_button.clicked.connect(self.next_click)
        next_button.setIcon(QIcon('.\\img\\down-arrow.png'))
        next_button.setIconSize(QSize(600,100))
        next_button.setShortcut('x')
        vLayout.addWidget(next_button) 

        buttonGroupBox.setLayout(vLayout)

        return buttonGroupBox

    def createHButtons(self):
        hBottensWidget = QWidget()
        hLayout = QHBoxLayout()

        back_button = QPushButton('a', self)
        back_button.clicked.connect(self.back_click)
        back_button.setIcon(QIcon('.\\img\\left-arrow.png'))
        back_button.setIconSize(QSize(200,150))
        back_button.setShortcut('a')
        hLayout.addWidget(back_button) 
 
        cur_button = QPushButton('s', self)
        cur_button.clicked.connect(self.cur_click)
        cur_button.setIcon(QIcon('.\\img\\reload.png'))
        cur_button.setIconSize(QSize(200,150))
        cur_button.setShortcut('s')
        hLayout.addWidget(cur_button) 
 
        forward_button = QPushButton('d', self)
        forward_button.clicked.connect(self.forward_click)
        forward_button.setIcon(QIcon('.\\img\\right-arrow.png'))
        forward_button.setIconSize(QSize(200,150))
        forward_button.setShortcut('d')
        hLayout.addWidget(forward_button) 
 
        hBottensWidget.setLayout(hLayout)

        return hBottensWidget

    def createHorizontalButtons2(self):
        self.horizontalButtons2 = QGroupBox("설정 변경")
        layout = QHBoxLayout()
 
        plusRowButton = QPushButton('행+', self)
        plusRowButton.clicked.connect(self.plus_row)
        layout.addWidget(plusRowButton) 
 
        plusColButton = QPushButton('열+', self)
        plusColButton.clicked.connect(self.plus_col)
        layout.addWidget(plusColButton) 

        init_button = QPushButton('초기화', self)
        init_button.clicked.connect(self.initialize)
        layout.addWidget(init_button) 
 
        self.horizontalButtons2.setLayout(layout)
 
    def createLogTable(self):
        # Create table
        self.logTable = QTableWidget()
        self.logTable.setRowCount(self.tableRow)
        self.logTable.setColumnCount(self.tableCol)
        self.logTable.move(0,0)

    def createConfigTable(self):
        self.configTable = QTableWidget()

        try:
            # load configurationFile
            cwb = load_workbook(self.configFile.strip())
        except:
            # message box
            string = "설정파일을 불러올 수 없습니다."
            QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)

        else:
            # Get matadata
            cws = cwb.active
            self.crow = cws.cell(row = 1, column = 1).value
            self.ccol = cws.cell(row = 1, column = 2).value

            # Configure table
            self.configTable.setRowCount(self.crow)
            self.configTable.setColumnCount(self.ccol)
            self.configTable.move(0,0)

            # Load data from configFile
            for i in range(self.crow):
                for j in range(self.ccol):
                    item = str(cws.cell(row = i+2, column = j+1).value).strip()
                    if item == 'None':
                        item = ''
                    self.configTable.setItem(i, j, QTableWidgetItem(item))

            # check if files exist
            arr = listdir('./audio_clips')
            for row in range(self.crow):
                for col in range(self.ccol):
                    # if 박스(0,0)
                    if row == 0 and col == 0:
                        continue;

                    # reset backgound color
                    self.configTable.item(row, col).setBackground(QColor(255,0,0))
                    
                    # if file exist, change background color
                    fname = str(row) + '_' + str(col) + '.mp3'
                    if fname in arr:
                        if row == 0:
                            self.configTable.item(row, col).setBackground(QColor(255,255,0))
                        else:
                            self.configTable.item(row, col).setBackground(QColor(0,255,255))

            # 박스(0,0)
            self.configTable.setItem(0, 0, QTableWidgetItem('박스'))
            self.configTable.item(0, 0).setBackground(QColor(255,255,255))
            
            # link the callback function            
            self.configTable.itemDoubleClicked.connect(self.item_doubleClicked)
            self.configTable.itemChanged.connect(self.item_changed)


    def setLogTable(self):
        # shifting other rows
        for r in range(1, self.tableRow):
            for c in range(self.tableCol):
                try:
                    self.logTable.item(r,c).setBackground(QColor(255,255,255))
                    self.logTable.setItem(r-1 , c, self.logTable.item(r,c).clone())
                except:
                    pass

        # set current row
        for idx, key in enumerate(list(self.dict.keys())):
            if type(self.dict[key]) is list:
                self.logTable.setItem(self.tableRow - 1, idx, QTableWidgetItem(' '.join(self.dict[key])))
                self.logTable.item(self.tableRow - 1, idx).setBackground(QColor(255,255,0))
            else:
                self.logTable.setItem(self.tableRow - 1, idx, QTableWidgetItem(self.dict[key]))
                self.logTable.item(self.tableRow - 1, idx).setBackground(QColor(255,255,0))

    def read(self):

        # 포장 순번
        self.dict['num'] = str(self.ws.cell(row = self.row, column = 3).value[2:]).strip()

        # 거래처명
        self.dict['partner'] = str(self.ws.cell(row = self.row, column = 5).value).strip()

        # 배송센터
        self.dict['parcel'] = str(self.ws.cell(row = self.row, column = 4).value).strip()

        # 특이사항
        self.dict['exception'] = str(self.ws.cell(row = self.row, column = 6).value).strip()

        # 박스
        self.dict['box'] = str(self.ws.cell(row = self.row, column = 2).value).strip()

        # left things
        print(len(self.dict))
        for i in range(5, len(self.dict)):
            header = str(self.ws.cell(row = 1, column = i + 2).value).strip()
            self.dict[header] = str(self.ws.cell(row = self.row, column = i + 2).value).strip()
            self.parsing(header, self.dict[header])

        print(self.dict)
        self.setLogTable()


    def parsing(self, key, val):
        if val == 'None' or val == '':
            self.dict[key] = None

        else:
            if '(' in val or '/' in val:
                arr = re.split('[(/]',val)
                for i in range(len(arr)):
                    if ')' in arr[i]:
                        arr[i] = arr[i][:arr[i].index(')')]
                    arr[i] = arr[i].strip()
                self.dict[key] = arr

            else:
                self.dict[key] = val

    def itemFromKeyVal(self, key, val):
        items = self.configTable.findItems(val, Qt.MatchExactly)
        if len(items) <= 0:
            # Error
            string = '('+key+', '+val+') '+'아이템을 찾을 수 없습니다.'
            QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)
        else:
            for item in items:
                if self.configTable.item(0, item.column()).data(0) == key:
                    return item


    def load_audiolist(self):
        for key, val in self.dict.items():

            # 포장순번     
            if key == 'num':
                for i in range(len(self.dict['num'])):
                    self.audiolist.append('_' + val[i])

                # beep
                self.audiolist.append('_beep')

            # 택배발송
            elif key == 'parcel':
                if val == '택배발송':
                    self.audiolist.append('_택배발송')

                    # beep
                    self.audiolist.append('_beep')
            
            # 박스
            elif key == 'box':
                item = self.itemFromKeyVal('박스', val)
                if item:
                    self.audiolist.append(str(item.row()) + '_' + str(item.column()))

                    # beep
                    self.audiolist.append('_beep')

            elif key in ['partner', 'exception']:
                pass
            
            # general case
            else:
                # The case(val == None) will be ignored 
                if val == None:
                    pass
                
                # when val is list
                elif type(val) == list:
                    for idx, eachVal in enumerate(val):
                        item = self.itemFromKeyVal(key, eachVal)
                        if item:
                            if idx == 0:
                                self.audiolist.append('0_' + str(item.column())) # key
                            self.audiolist.append(str(item.row()) + '_' + str(item.column())) # val

                    # beep
                    self.audiolist.append('_beep')

                # when val is not list
                else:
                    item = self.itemFromKeyVal(key, val)
                    if item:
                        if val == '1' or key == val:
                            self.audiolist.append('0_' + str(item.column())) # key
                        else:
                            self.audiolist.append('0_' + str(item.column())) # key
                            self.audiolist.append(str(item.row()) + '_' + str(item.column())) # val

                        # beep
                        self.audiolist.append('_beep')

            

        print(self.audiolist)

    def speak(self):
        self.playlist.clear()
        for clip in self.audiolist:
            url = QUrl.fromLocalFile('./audio_clips/' + clip + '.mp3')
            #print(url)
            self.playlist.addMedia(QMediaContent(url))
        self.player.setPlaylist(self.playlist)
        self.player.play()
            
#----------------------- Control button callback -----------------------#

    @pyqtSlot()
    def pre_click(self):
        if not self.fileReady:
            self.cur_click()
        
        else:
            if self.row == 2:
                self.statusbar.showMessage("Can't be previous.")
            else:
                self.row -= 1

            self.dict = self.dict.fromkeys(self.dict, None)
            del self.audiolist[:]
            self.read()
            self.load_audiolist()
            self.speak()


    @pyqtSlot()
    def cur_click(self):
        if not self.fileReady:
            string = '파일이 준비되어 있지 않습니다.'
            QMessageBox.question(self, '경고', string, QMessageBox.Ok, QMessageBox.Ok)
            self.openFileNameDialog()
        else:

            self.dict = self.dict.fromkeys(self.dict, None)
            del self.audiolist[:]
            self.read()
            self.load_audiolist()
            self.speak()


    @pyqtSlot()
    def next_click(self):
        if not self.fileReady:
            self.cur_click()

        else:
            if self.row == self.ws.max_row:
                self.statusbar.showMessage("It's over.")
            else:
                self.row += 1

            self.dict = self.dict.fromkeys(self.dict, None)
            del self.audiolist[:]
            self.read()
            self.load_audiolist()
            self.speak()

    @pyqtSlot()
    def back_click(self):
        if self.playlist.mediaCount() == 0:
            self.cur_click()
        elif self.playlist.mediaCount() != 0:
            p = re.compile('.+_beep.+')
            cnt = 0
            for i in range(self.playlist.mediaCount()):
                # if it's start point, start at here
                if self.playlist.currentIndex() == 0:
                    break
                # go backward
                self.playlist.setCurrentIndex(self.playlist.previousIndex(1))
                
                # start at previous beep point
                if p.match(str(self.playlist.currentMedia().canonicalUrl())):
                    cnt = cnt + 1
                    if cnt == 2:
                        print(self.playlist.currentIndex())
                        if self.player.state() == QMediaPlayer.StoppedState:
                            self.player.play()
                        break
                


    @pyqtSlot()
    def forward_click(self):
        if self.playlist.mediaCount() == 0:
            self.cur_click()
        elif self.playlist.mediaCount() != 0:
             p = re.compile('.+_beep.+')
             for i in range(self.playlist.mediaCount()):
                # don't go further from end point
                if self.playlist.currentIndex() < 0:
                    break

                # go forward
                self.playlist.setCurrentIndex(self.playlist.nextIndex(1))
                
                # start at next beep point
                if p.match(str(self.playlist.currentMedia().canonicalUrl())):
                    print(self.playlist.currentIndex())
                    break


#----------------------- Configuration button callback -----------------------#

    @pyqtSlot()
    def plus_row(self):
        # change configTable
        self.crow = self.crow + 1
        self.configTable.setRowCount(self.crow)
        
        # fill the generated cell
        self.configTable.itemChanged.disconnect(self.item_changed)
        for col in range(self.ccol):
            self.configTable.setItem(self.crow - 1, col, QTableWidgetItem(''))
            self.configTable.item(self.crow - 1, col).setBackground(QColor(255,0,0))
        self.configTable.itemChanged.connect(self.item_changed)



    @pyqtSlot()
    def plus_col(self):
        # change configTable
        self.ccol = self.ccol + 1
        self.configTable.setColumnCount(self.ccol)

        # fill the generated cell
        self.configTable.itemChanged.disconnect(self.item_changed)
        for row in range(self.crow):
            self.configTable.setItem(row, self.ccol - 1, QTableWidgetItem(''))
            self.configTable.item(row, self.ccol - 1).setBackground(QColor(255,0,0))
        self.configTable.itemChanged.connect(self.item_changed)


    @pyqtSlot()
    def initialize(self):
        # remove configurable audio files
        arr = listdir('./audio_clips')
        p = re.compile('_.+')
        for file in arr:
            if p.match(file):
                continue
            remove('./audio_clips/' + file)

        # init configTable
        self.configTable.itemChanged.disconnect(self.item_changed) # lock

        self.configTable.clear()
        self.crow = 5
        self.ccol = 5
        self.configTable.setRowCount(self.crow)
        self.configTable.setColumnCount(self.ccol)

        # reset configTable item
        for row in range(self.crow):
            for col in range(self.ccol):
                self.configTable.setItem(row, col, QTableWidgetItem(''))
                self.configTable.item(row, col).setBackground(QColor(255,0,0))
        self.configTable.setItem(0, 0, QTableWidgetItem('박스'))
        self.configTable.item(0, 0).setBackground(QColor(255,255,255))

        self.configTable.itemChanged.connect(self.item_changed) # unlock

        # init configFile
        self.update_configFile()

#---------------------- ConfigTable signal callback ------------------------#

    def item_doubleClicked(self,item):
        self.previousItem = item.data(0)
        print(self.previousItem)

    def update_configFile(self):
        cwb_w = Workbook(write_only=True)
        cws_w = cwb_w.create_sheet()

        cws_w.append([self.crow, self.ccol])

        for row in range(self.crow):
            itemList = []
            for col in range(self.ccol):
                itemList.append(self.configTable.item(row, col).data(0))
            cws_w.append(itemList)
        try:
            cwb_w.save(self.configFile)
        except:
                string = """설정파일이 열려있을 수 있습니다.
                            설정파일을 닫은 후에 다시 시도하세요."""
                QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)

    def item_changed(self,item):
        self.configTable.itemChanged.disconnect(self.item_changed) # lock
        
        if item.row() == 0 and item.column() == 0:
            string = "이 항목은 바꿀 수 없습니다."
            QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)
            self.configTable.setItem(item.row(), item.column(), QTableWidgetItem(self.previousItem))

        else:
            # get file name 
            fileName, _ = QFileDialog.getOpenFileName(self,"Open file", "", "All Files (*)")

            # count the number of key that has same name
            keys = self.configTable.findItems(item.data(0), Qt.MatchExactly)
            kcnt = 0
            for key in keys:
                if key.row() == 0:
                    kcnt = kcnt + 1

            # count the number of atribute that has same name
            atributes = self.configTable.findItems(item.data(0), Qt.MatchExactly)
            acnt = 0
            for atribute in atributes:
                if atribute.row() == 0:
                    pass
                elif atribute.column() == item.column():
                    acnt = acnt + 1

            # change is accepted only in case of uniqueness and existence and it is not 박스(0,0)
            if kcnt >= 2:
                string = "항목명이 같을 수 없습니다."
                QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)
                self.configTable.setItem(item.row(), item.column(), QTableWidgetItem(self.previousItem))

            elif acnt >= 2:
                string = "같은 항목에 같은 이름의 아이템을 둘 수 없습니다."
                QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)
                self.configTable.setItem(item.row(), item.column(), QTableWidgetItem(self.previousItem))

            elif fileName:
                # copy file to local dir
                dst = "./audio_clips./" + str(item.row()) + '_' + str(item.column()) + '.mp3'
                copyfile(fileName, dst)

                # change cell color
                if item.row() == 0:
                    self.configTable.item(item.row(), item.column()).setBackground(QColor(255,255,0))
                else:
                    self.configTable.item(item.row(), item.column()).setBackground(QColor(0,255,255))

                # update configFile
                self.update_configFile()
                
            else:
                string = "선택이 취소됨."
                QMessageBox.question(self, 'Error', string, QMessageBox.Ok, QMessageBox.Ok)
                self.configTable.setItem(item.row(), item.column(), QTableWidgetItem(self.previousItem))


        self.configTable.itemChanged.connect(self.item_changed) # unlock



 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())